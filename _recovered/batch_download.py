"""Excel/CSV 批量解析、标准匹配与 ZIP 打包下载。"""
from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from core.db import db
from core.std_normalize import normalize_std_id

MAX_ROWS = 150
DISK_TIMEOUT = 12

_STD_HEADER_KEYS = (
    "标准编号",
    "标准号",
    "编号",
    "std_id",
    "stdid",
    "标准代码",
    "标准代号",
    "标准文号",
)
_NAME_HEADER_KEYS = (
    "标准名称",
    "名称",
    "标准名",
    "title",
    "中文名称",
    "标准中文名称",
    "名称关键词",
)


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _header_matches(value: Any, keys: tuple[str, ...], *substrings: str) -> bool:
    nh = _norm_header(value)
    if not nh:
        return False
    if nh in {_norm_header(k) for k in keys}:
        return True
    return any(s in nh for s in substrings)


def detect_columns(headers: list[Any]) -> dict[str, int | None]:
    std_col: int | None = None
    name_col: int | None = None
    for i, h in enumerate(headers):
        if _header_matches(h, _STD_HEADER_KEYS, "编号", "标准号", "std"):
            std_col = i
        if _header_matches(h, _NAME_HEADER_KEYS, "名称", "title"):
            name_col = i
    return {"std_col": std_col, "name_col": name_col}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _row_query(cells: list[str], cols: dict[str, int | None]) -> str:
    std_col = cols.get("std_col")
    name_col = cols.get("name_col")
    std_val = cells[std_col] if std_col is not None and std_col < len(cells) else ""
    name_val = cells[name_col] if name_col is not None and name_col < len(cells) else ""
    if std_val:
        return std_val
    if name_val:
        return name_val
    for c in cells:
        if c:
            return c
    return ""


def _looks_like_header_row(cells: list[str]) -> bool:
    joined = " ".join(cells).casefold()
    if not joined:
        return False
    hints = ("标准", "编号", "名称", "std", "title", "序号", "行号")
    return sum(1 for h in hints if h in joined) >= 2


def _parse_csv(data: bytes) -> tuple[list[dict], dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    raw_rows = [[_cell_text(c) for c in row] for row in reader]
    return _rows_from_matrix(raw_rows, source="csv")


def _parse_xlsx(data: bytes) -> tuple[list[dict], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请运行 pip install openpyxl") from exc

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_text(c) for c in row]
            if any(cells):
                raw_rows.append(cells)
    finally:
        wb.close()
    return _rows_from_matrix(raw_rows, source="xlsx")


def _rows_from_matrix(
    raw_rows: list[list[str]], *, source: str
) -> tuple[list[dict], dict[str, Any]]:
    if not raw_rows:
        return [], {"source": source, "columns": {}, "header_row": None, "total_rows": 0}

    header_row_idx = 0
    cols = detect_columns(raw_rows[0])
    if cols["std_col"] is None and cols["name_col"] is None:
        if _looks_like_header_row(raw_rows[0]):
            header_row_idx = 0
        else:
            header_row_idx = -1
            cols = {"std_col": 0, "name_col": None}
    else:
        header_row_idx = 0

    items: list[dict] = []
    start = header_row_idx + 1 if header_row_idx >= 0 else 0
    for i in range(start, len(raw_rows)):
        cells = raw_rows[i]
        query = _row_query(cells, cols)
        if not query:
            continue
        items.append(
            {
                "row": i + 1,
                "query": query,
                "std_hint": cells[cols["std_col"]] if cols.get("std_col") is not None and cols["std_col"] < len(cells) else "",
                "name_hint": cells[cols["name_col"]] if cols.get("name_col") is not None and cols["name_col"] < len(cells) else "",
            }
        )
        if len(items) >= MAX_ROWS:
            break

    meta = {
        "source": source,
        "columns": cols,
        "header_row": header_row_idx + 1 if header_row_idx >= 0 else None,
        "total_rows": len(items),
        "truncated": len(raw_rows) - start > len(items) and len(items) >= MAX_ROWS,
        "max_rows": MAX_ROWS,
    }
    return items, meta


def parse_upload(filename: str, data: bytes) -> dict[str, Any]:
    ext = Path(filename or "").suffix.casefold()
    if ext == ".csv":
        items, meta = _parse_csv(data)
    elif ext in (".xlsx", ".xlsm"):
        items, meta = _parse_xlsx(data)
    else:
        return {
            "ok": False,
            "error": "仅支持 .xlsx、.xlsm 或 .csv 文件",
        }
    if not items:
        return {
            "ok": False,
            "error": "未识别到有效查询行，请确认 Excel 含有「标准编号」或「标准名称」列",
            "meta": meta,
        }
    return {"ok": True, "items": items, "meta": meta}


def _collect_files_for_standard(std, *, scan_disk: bool, disk_timeout: int):
    from backend.app import _collect_files_for_standard

    return _collect_files_for_standard(
        std,
        scan_disk=scan_disk,
        disk_timeout=disk_timeout,
        deep=scan_disk,
    )


def _find_pdf_on_disk(rel_path: str, file_name: str, std_id: str | None = None):
    from backend.app import _find_pdf_on_disk

    return _find_pdf_on_disk(rel_path, file_name, std_id=std_id, deep=True, disk_timeout=DISK_TIMEOUT)


def _pick_pdf_path(std, files: list[dict]) -> Path | None:
    for f in files:
        if not f.get("exists"):
            continue
        if f.get("source") == "disk":
            rel = f.get("file_path") or ""
            if rel:
                candidate = Path(rel)
                if candidate.is_file():
                    return candidate
            continue
        found = _find_pdf_on_disk(
            f.get("file_path") or "",
            f.get("file_name") or "",
            std_id=std.std_id,
        )
        if found:
            return found
    return None


def resolve_item(query: str, *, scan_disk: bool = True) -> dict[str, Any]:
    q = (query or "").strip()
    if not q:
        return {"status": "empty", "query": q, "message": "空行"}

    if not db.is_ready():
        return {"status": "error", "query": q, "message": "标准库未就绪"}

    hits = db.search(q, limit=5)
    if not hits:
        return {"status": "not_found", "query": q, "message": "未找到匹配标准"}

    std = hits[0]
    files = _collect_files_for_standard(std, scan_disk=scan_disk, disk_timeout=DISK_TIMEOUT)
    pdf_path = _pick_pdf_path(std, files)
    if not pdf_path:
        return {
            "status": "no_pdf",
            "query": q,
            "std_id": std.std_id,
            "std_chinesename": std.std_chinesename,
            "base_id": std.id,
            "message": "已匹配标准但未找到 PDF",
        }

    zip_name = _safe_zip_name(std.std_id, pdf_path.name)
    return {
        "status": "ok",
        "query": q,
        "std_id": std.std_id,
        "std_chinesename": std.std_chinesename,
        "base_id": std.id,
        "file_name": pdf_path.name,
        "zip_name": zip_name,
        "pdf_path": str(pdf_path),
        "file_size": pdf_path.stat().st_size if pdf_path.is_file() else None,
    }


def _safe_zip_name(std_id: str, original: str) -> str:
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", std_id or "unknown")
    base = base.replace(" ", "")[:80] or "unknown"
    suffix = Path(original).suffix or ".pdf"
    stem = Path(original).stem
    stem_safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", stem)[:60]
    return f"{base}_{stem_safe}{suffix}"


def _unique_name(used: set[str], name: str) -> str:
    if name not in used:
        used.add(name)
        return name
    stem = Path(name).stem
    suffix = Path(name).suffix
    n = 2
    while True:
        candidate = f"{stem}_{n}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        n += 1


def build_zip_archive(
    items: list[dict],
    *,
    scan_disk: bool = True,
    progress: Callable[[int, int], None] | None = None,
) -> tuple[io.BytesIO, dict[str, Any]]:
    buf = io.BytesIO()
    used_names: set[str] = set()
    results: list[dict] = []
    ok_count = 0

    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        total = len(items)
        for idx, item in enumerate(items, start=1):
            if progress:
                progress(idx, total)
            query = (item.get("query") or "").strip()
            row_no = item.get("row") or idx
            resolved = resolve_item(query, scan_disk=scan_disk)
            resolved["row"] = row_no
            results.append(resolved)
            if resolved.get("status") != "ok":
                continue
            pdf_path = Path(resolved["pdf_path"])
            if not pdf_path.is_file():
                resolved["status"] = "no_pdf"
                resolved["message"] = "PDF 文件不存在"
                continue
            prefix = f"{row_no:03d}_"
            entry_name = _unique_name(used_names, prefix + resolved["zip_name"])
            zf.write(pdf_path, arcname=entry_name)
            resolved["zip_entry"] = entry_name
            ok_count += 1

        manifest = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "total": total,
            "success": ok_count,
            "failed": total - ok_count,
            "results": results,
        }
        zf.writestr(
            "_批量下载清单.json",
            json.dumps(manifest, ensure_ascii=False, indent=2),
        )
        lines = [
            "标准 PDF 批量下载清单",
            f"生成时间：{manifest['generated_at']}",
            f"共 {total} 条，成功 {ok_count} 条，失败 {total - ok_count} 条",
            "",
        ]
        for r in results:
            mark = "✓" if r.get("status") == "ok" else "✗"
            std_id = r.get("std_id") or "—"
            msg = r.get("message") or r.get("status") or ""
            lines.append(f"{mark} 第{r.get('row')}行 | 查询：{r.get('query')} | {std_id} | {msg}")
        zf.writestr("_批量下载清单.txt", "\n".join(lines))

    buf.seek(0)
    summary = {
        "total": len(items),
        "success": ok_count,
        "failed": len(items) - ok_count,
        "results": results,
    }
    return buf, summary


def preview_items(items: list[dict], *, scan_disk: bool = False) -> dict[str, Any]:
    """快速预览：默认不扫磁盘，仅查库内索引。"""
    rows: list[dict] = []
    for item in items[:MAX_ROWS]:
        q = (item.get("query") or "").strip()
        if not q:
            continue
        rows.append(resolve_item(q, scan_disk=scan_disk))
    ok = sum(1 for r in rows if r.get("status") == "ok")
    return {
        "ok": True,
        "items": rows,
        "summary": {"total": len(rows), "success": ok, "failed": len(rows) - ok},
    }
