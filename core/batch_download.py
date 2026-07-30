"""Excel/CSV 批量解析、标准匹配与 ZIP 打包下载。"""
from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from core.db import db
from core.pdf_service import collect_files_for_standard, pick_pdf_path
from core.std_normalize import normalize_std_id

MAX_ROWS = 400

_STD_HEADER_KEYS = (
    "标准编号",
    "标准号",
    "编号",
    "std_id",
    "stdid",
    "标准代码",
    "标准代号",
    "标准文号",
)
_NAME_HEADER_KEYS = (
    "标准名称",
    "名称",
    "标准名",
    "title",
    "中文名称",
    "标准中文名称",
    "名称关键词",
)
_REMARK_HEADER_KEYS = ("备注", "说明", "remark", "note")
_STD_NUMBER_RE = re.compile(
    r"^(?:GB/?T?|GB/Z|GB|GJB/?T?|DL/?T?|DL|JB/?T?|JB|JJF|JJG|HG/?T?|NY/?T?|SH/?T?|SY/?T?|ISO|IEC|ASTM|EN|Q(?:/|[\s/]))",
    re.I,
)


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _header_matches(value: Any, keys: tuple[str, ...], *substrings: str) -> bool:
    nh = _norm_header(value)
    if not nh:
        return False
    if nh in {_norm_header(k) for k in keys}:
        return True
    return any(s in nh for s in substrings)


def detect_columns(headers: list[Any]) -> dict[str, int | None]:
    std_col: int | None = None
    name_col: int | None = None
    remark_col: int | None = None
    for i, h in enumerate(headers):
        if _header_matches(h, _STD_HEADER_KEYS, "标准号", "编号", "std"):
            std_col = i
        if _header_matches(h, _NAME_HEADER_KEYS, "名称", "title"):
            name_col = i
        if _header_matches(h, _REMARK_HEADER_KEYS, "备注"):
            remark_col = i
    return {"std_col": std_col, "name_col": name_col, "remark_col": remark_col}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def looks_like_std_number(text: str) -> bool:
    s = (text or "").strip()
    if len(s) < 3:
        return False
    if _STD_NUMBER_RE.match(s):
        return True
    if re.match(r"^GB", s, re.I) and re.search(r"\d", s):
        return True
    if s.startswith("国网") or s.startswith("Q/"):
        return True
    if re.search(r"[\(（][^)）]+[\)）]", s) and re.search(r"\d", s):
        return True
    compact = re.sub(r"\s+", "", s)
    return bool(re.search(r"\d{2,}", compact) and re.search(r"[A-Za-z/（）()\-—／]", compact))


def _find_header_row(raw_rows: list[list[str]]) -> tuple[int, dict[str, int | None]] | None:
    for i in range(min(len(raw_rows), 25)):
        cols = detect_columns(raw_rows[i])
        if cols["std_col"] is not None:
            return i, cols
    return None


def _should_skip_row(query: str) -> bool:
    q = (query or "").strip()
    if not q:
        return True
    if re.fullmatch(r"\d{1,4}", q):
        return True
    if "明细表" in q or "标准清单" in q:
        return True
    return not looks_like_std_number(q)


def _rows_from_matrix(
    raw_rows: list[list[str]], *, source: str
) -> tuple[list[dict], dict[str, Any]]:
    if not raw_rows:
        return [], {"source": source, "columns": {}, "header_row": None, "total_rows": 0}

    header_row_idx = -1
    cols: dict[str, int | None] = {"std_col": None, "name_col": None, "remark_col": None}
    found = _find_header_row(raw_rows)
    if found:
        header_row_idx, cols = found
    else:
        for i in range(min(len(raw_rows), 40)):
            for j, cell in enumerate(raw_rows[i]):
                if looks_like_std_number(cell):
                    cols = {"std_col": j, "name_col": None, "remark_col": None}
                    header_row_idx = i - 1 if i > 0 else -1
                    break
            if cols["std_col"] is not None:
                break

    if cols["std_col"] is None:
        return [], {
            "source": source,
            "columns": cols,
            "header_row": None,
            "total_rows": 0,
        }

    items: list[dict] = []
    start = header_row_idx + 1 if header_row_idx >= 0 else 0
    for i in range(start, len(raw_rows)):
        cells = raw_rows[i]
        std_col = cols["std_col"]
        query = cells[std_col] if std_col is not None and std_col < len(cells) else ""
        query = _cell_text(query)
        if _should_skip_row(query):
            continue
        name_col = cols.get("name_col")
        items.append(
            {
                "row": i + 1,
                "query": query,
                "std_hint": query,
                "name_hint": cells[name_col]
                if name_col is not None and name_col < len(cells)
                else "",
            }
        )
        if len(items) >= MAX_ROWS:
            break

    meta = {
        "source": source,
        "columns": cols,
        "header_row": header_row_idx + 1 if header_row_idx >= 0 else None,
        "total_rows": len(items),
        "truncated": len(raw_rows) - start > len(items) and len(items) >= MAX_ROWS,
        "max_rows": MAX_ROWS,
    }
    return items, meta


def _parse_csv(data: bytes) -> tuple[list[dict], dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    raw_rows = [[_cell_text(c) for c in row] for row in reader]
    return _rows_from_matrix(raw_rows, source="csv")


def _parse_xlsx(data: bytes) -> tuple[list[dict], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请运行 pip install openpyxl") from exc

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_text(c) for c in row]
            if any(cells):
                raw_rows.append(cells)
    finally:
        wb.close()
    return _rows_from_matrix(raw_rows, source="xlsx")


def parse_upload(filename: str, data: bytes) -> dict[str, Any]:
    ext = Path(filename or "").suffix.casefold()
    if ext == ".csv":
        items, meta = _parse_csv(data)
    elif ext in (".xlsx", ".xlsm"):
        items, meta = _parse_xlsx(data)
    else:
        return {
            "ok": False,
            "error": "仅支持 .xlsx、.xlsm 或 .csv 文件",
        }
    if not items:
        return {
            "ok": False,
            "error": "未识别到有效标准号，请确认含有「标准号/标准编号」列",
            "meta": meta,
        }
    return {"ok": True, "items": items, "meta": meta}


def resolve_item(query: str) -> dict[str, Any]:
    q = (query or "").strip()
    if not q:
        return {"status": "empty", "query": q, "message": "空行"}

    if not db.is_ready():
        return {"status": "error", "query": q, "message": "标准库未就绪，请配置 MySQL 连接"}

    std = db.search_std_id(q)
    if not std:
        hits = db.search(q, limit=3)
        if hits and normalize_std_id(hits[0].std_id) == normalize_std_id(q):
            std = hits[0]
    if not std:
        return {
            "status": "not_found",
            "query": q,
            "std_id": q,
            "std_chinesename": "（库中未找到）",
            "message": "未找到匹配标准",
            "match_status": "not_found",
        }

    if _is_abolished(std):
        return {
            "status": "abolished",
            "query": q,
            "std_id": std.std_id,
            "std_chinesename": std.std_chinesename,
            "base_id": std.id,
            "ex_state": 0,
            "ex_state_label": "废止",
            "message": "废止标准，已跳过下载",
            "match_status": "abolished",
        }

    files = collect_files_for_standard(std)
    pdf_path = pick_pdf_path(std, files)
    if not pdf_path:
        return {
            "status": "no_pdf",
            "query": q,
            "std_id": std.std_id,
            "std_chinesename": std.std_chinesename,
            "base_id": std.id,
            "message": "有标准条目，但无 PDF",
            "match_status": "no_pdf",
        }

    zip_name = _safe_zip_name(std.std_id, pdf_path.name)
    return {
        "status": "ok",
        "query": q,
        "std_id": std.std_id,
        "std_chinesename": std.std_chinesename,
        "base_id": std.id,
        "file_name": pdf_path.name,
        "pdf_name": pdf_path.name,
        "zip_name": zip_name,
        "pdf_path": str(pdf_path),
        "file_size": pdf_path.stat().st_size if pdf_path.is_file() else None,
    }


def _safe_zip_name(std_id: str, original: str) -> str:
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", std_id or "unknown")
    base = base.replace(" ", "")[:80] or "unknown"
    suffix = Path(original).suffix or ".pdf"
    stem = Path(original).stem
    stem_safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", stem)[:60]
    return f"{base}_{stem_safe}{suffix}"


def _unique_name(used: set[str], name: str) -> str:
    if name not in used:
        used.add(name)
        return name
    stem = Path(name).stem
    suffix = Path(name).suffix
    n = 2
    while True:
        candidate = f"{stem}_{n}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        n += 1


def _is_abolished(std) -> bool:
    """ex_state=0 为废止；兼容状态文案。"""
    if getattr(std, "ex_state", None) == 0:
        return True
    label = (getattr(std, "ex_state_label", None) or getattr(std, "std_status", None) or "").strip()
    return label in ("废止", "作废")


def _excel_remark(status: str | None) -> str:
    if status == "ok":
        return ""
    if status == "abolished":
        return "废止"
    return "否"


def _failed_status(status: str | None) -> bool:
    return status in ("not_found", "no_pdf", "error", "empty", "abolished")


def _abolished_list_text(results: list[dict], *, generated_at: str) -> str:
    abolished = [r for r in results if r.get("status") == "abolished"]
    lines = [
        "跳过废止标准名单",
        f"生成时间：{generated_at}",
        f"共 {len(abolished)} 条废止标准未下载",
        "",
        "说明：状态为「废止」的标准文件不纳入本次下载。",
        "",
    ]
    if not abolished:
        lines.append("（本次无废止标准被跳过）")
    else:
        for r in abolished:
            row = r.get("row")
            row_part = f"第{row}行 | " if row is not None else ""
            lines.append(
                f"- {row_part}标准号：{r.get('std_id') or r.get('query') or '—'} | "
                f"名称：{r.get('std_chinesename') or '—'} | 原因：废止，已跳过"
            )
    return "\n".join(lines)


def build_annotated_excel(
    original_data: bytes,
    filename: str,
    results: list[dict],
    meta: dict[str, Any] | None,
) -> tuple[bytes, str]:
    """在原 Excel/CSV 的备注列标注「否」，返回 (文件字节, 压缩包内文件名)。"""
    meta = meta or {}
    by_row = {int(r["row"]): r for r in results if r.get("row") is not None}
    ext = Path(filename or "").suffix.casefold()
    stem = Path(filename or "批量清单").stem or "批量清单"
    out_name = f"{stem}_下载结果.xlsx"

    if ext == ".csv":
        return _build_annotated_csv(original_data, by_row, meta, stem)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，请运行 pip install openpyxl") from exc

    wb = load_workbook(filename=io.BytesIO(original_data))
    ws = wb.active
    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        raw_rows.append([_cell_text(c) for c in row])

    header_idx = (meta.get("header_row") or 1) - 1
    if header_idx < 0 or header_idx >= len(raw_rows):
        found = _find_header_row(raw_rows)
        header_idx = found[0] if found else 0

    cols = detect_columns(raw_rows[header_idx]) if header_idx < len(raw_rows) else {}
    remark_col = cols.get("remark_col")
    if remark_col is None:
        remark_col = len(raw_rows[header_idx]) if header_idx < len(raw_rows) else 0
        if header_idx < len(raw_rows):
            while len(raw_rows[header_idx]) <= remark_col:
                raw_rows[header_idx].append("")
            if not raw_rows[header_idx][remark_col]:
                raw_rows[header_idx][remark_col] = "备注"

    for i in range(header_idx + 1, len(raw_rows)):
        cells = raw_rows[i]
        std_col = cols.get("std_col")
        q = cells[std_col] if std_col is not None and std_col < len(cells) else ""
        if not looks_like_std_number(_cell_text(q)):
            continue
        while len(cells) <= remark_col:
            cells.append("")
        hit = by_row.get(i + 1)
        cells[remark_col] = _excel_remark(hit.get("status") if hit else None)
        for j, val in enumerate(cells):
            ws.cell(row=i + 1, column=j + 1, value=val or None)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), out_name


def _build_annotated_csv(
    original_data: bytes,
    by_row: dict[int, dict],
    meta: dict[str, Any],
    stem: str,
) -> tuple[bytes, str]:
    text = original_data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [[_cell_text(c) for c in row] for row in reader]
    found = _find_header_row(rows)
    header_idx = found[0] if found else 0
    cols = found[1] if found else detect_columns(rows[header_idx])
    remark_col = cols.get("remark_col")
    if remark_col is None:
        remark_col = len(rows[header_idx])
        rows[header_idx].append("备注")
    for i in range(header_idx + 1, len(rows)):
        std_col = cols.get("std_col")
        q = rows[i][std_col] if std_col is not None and std_col < len(rows[i]) else ""
        if not looks_like_std_number(_cell_text(q)):
            continue
        while len(rows[i]) <= remark_col:
            rows[i].append("")
        hit = by_row.get(i + 1)
        rows[i][remark_col] = _excel_remark(hit.get("status") if hit else None)
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerows(rows)
    return out.getvalue().encode("utf-8-sig"), f"{stem}_下载结果.csv"


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "标准清单"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="1E40AF")
    body_font = Font(name="Microsoft YaHei", size=10, color="334155")
    header_fill = PatternFill("solid", fgColor="EFF6FF")

    headers = ["标准号", "标准名", "备注"]
    ws.append(headers)
    ws.append(["GB/T 1002-2024", "单相插头插座", ""])
    ws.append(["GB 5749-2022", "生活饮用水卫生标准", ""])

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.border = border

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zip_archive(
    items: list[dict],
    *,
    progress: Callable[[int, int], None] | None = None,
    original_data: bytes | None = None,
    original_filename: str | None = None,
    parse_meta: dict[str, Any] | None = None,
) -> tuple[io.BytesIO, dict[str, Any]]:
    buf = io.BytesIO()
    used_names: set[str] = set()
    results: list[dict] = []
    ok_count = 0
    abolished_count = 0

    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        total = len(items)
        for idx, item in enumerate(items, start=1):
            if progress:
                progress(idx, total)
            query = (item.get("query") or "").strip()
            row_no = item.get("row") or idx
            resolved = resolve_item(query)
            resolved["row"] = row_no
            results.append(resolved)
            if resolved.get("status") == "abolished":
                abolished_count += 1
                continue
            if resolved.get("status") != "ok":
                continue
            pdf_path = Path(resolved["pdf_path"])
            if not pdf_path.is_file():
                resolved["status"] = "no_pdf"
                resolved["message"] = "PDF 文件不存在"
                continue
            prefix = f"{row_no:03d}_"
            entry_name = _unique_name(used_names, prefix + resolved["zip_name"])
            try:
                zf.write(pdf_path, arcname=f"PDF/{entry_name}")
            except OSError as exc:
                resolved["status"] = "error"
                resolved["message"] = f"写入 ZIP 失败：{exc}"
                continue
            resolved["zip_entry"] = entry_name
            ok_count += 1

        if original_data and original_filename:
            try:
                xbytes, xname = build_annotated_excel(
                    original_data, original_filename, results, parse_meta
                )
                zf.writestr(xname, xbytes)
            except Exception:
                pass

        generated_at = datetime.now().isoformat(timespec="seconds")
        abolished_items = [r for r in results if r.get("status") == "abolished"]
        # 清单 JSON 勿把整份 results 里偶发的不可序列化对象带崩
        safe_results = json.loads(json.dumps(results, ensure_ascii=False, default=str))
        manifest = {
            "generated_at": generated_at,
            "total": total,
            "success": ok_count,
            "failed": total - ok_count,
            "abolished": abolished_count,
            "abolished_items": [
                {
                    "row": r.get("row"),
                    "query": r.get("query"),
                    "std_id": r.get("std_id"),
                    "std_chinesename": r.get("std_chinesename"),
                    "message": r.get("message"),
                }
                for r in abolished_items
            ],
            "results": safe_results,
        }
        zf.writestr(
            "_批量下载清单.json",
            json.dumps(manifest, ensure_ascii=False, indent=2, default=str),
        )
        lines = [
            "标准 PDF 批量下载清单",
            f"生成时间：{generated_at}",
            f"共 {total} 条，成功 {ok_count} 条，跳过废止 {abolished_count} 条，"
            f"其他失败 {total - ok_count - abolished_count} 条",
            "",
        ]
        for r in results:
            if r.get("status") == "ok":
                mark = "✓"
            elif r.get("status") == "abolished":
                mark = "⊘"
            else:
                mark = "✗"
            std_id = r.get("std_id") or "—"
            msg = r.get("message") or r.get("status") or ""
            lines.append(
                f"{mark} 第{r.get('row')}行 | 查询：{r.get('query')} | {std_id} | {msg}"
            )
        zf.writestr("_批量下载清单.txt", "\n".join(lines))
        # 单独一份废止跳过名单，便于核对
        zf.writestr(
            "_跳过废止清单.txt",
            _abolished_list_text(results, generated_at=generated_at),
        )

    buf.seek(0)
    summary = {
        "total": len(items),
        "success": ok_count,
        "failed": len(items) - ok_count,
        "abolished": abolished_count,
        "abolished_items": [
            {
                "row": r.get("row"),
                "query": r.get("query"),
                "std_id": r.get("std_id"),
                "std_chinesename": r.get("std_chinesename"),
                "message": r.get("message"),
            }
            for r in results
            if r.get("status") == "abolished"
        ],
        "results": results,
    }
    return buf, summary


MAX_BULK_IDS = 100


def build_zip_from_geo(
    filters,
    *,
    q: str = "",
    pdf_only: bool = True,
) -> tuple[io.BytesIO, dict[str, Any]]:
    """按省/市/县等地域条件批量打包 ZIP（基于 mydate 起草单位关系）。"""
    from core.geo_download import MAX_GEO_DOWNLOAD, list_geo_base_ids

    base_ids = list_geo_base_ids(filters, q=q, pdf_only=pdf_only, limit=MAX_GEO_DOWNLOAD)
    if not base_ids:
        return io.BytesIO(), {
            "total": 0,
            "success": 0,
            "failed": 0,
            "results": [],
            "region": "",
        }

    items: list[dict] = []
    for idx, bid in enumerate(base_ids, start=1):
        std = db.get_by_id(bid)
        if not std or not std.std_id:
            items.append({"row": idx, "query": str(bid), "base_id": bid})
            continue
        items.append({"row": idx, "query": std.std_id.strip(), "base_id": bid})

    buf, summary = build_zip_archive(items)
    summary["region"] = filters.province
    if filters.city:
        summary["region"] = f"{filters.province}_{filters.city}"
    summary["requested"] = len(base_ids)
    return buf, summary


def build_zip_from_base_ids(
    base_ids: list[int],
) -> tuple[io.BytesIO, dict[str, Any]]:
    """按标准库 base_id 列表打包 ZIP（检索结果多项下载）。"""
    unique: list[int] = []
    seen: set[int] = set()
    for raw in base_ids:
        try:
            bid = int(raw)
        except (TypeError, ValueError):
            continue
        if bid <= 0 or bid in seen:
            continue
        seen.add(bid)
        unique.append(bid)
        if len(unique) >= MAX_BULK_IDS:
            break

    items: list[dict] = []
    for idx, bid in enumerate(unique, start=1):
        std = db.get_by_id(bid)
        if not std or not std.std_id:
            items.append({"row": idx, "query": str(bid), "base_id": bid})
            continue
        items.append({"row": idx, "query": std.std_id.strip(), "base_id": bid})

    return build_zip_archive(items)


def preview_items(items: list[dict]) -> dict[str, Any]:
    rows: list[dict] = []
    for item in items[:MAX_ROWS]:
        q = (item.get("query") or "").strip()
        if not q:
            continue
        resolved = resolve_item(q)
        resolved["row"] = item.get("row")
        rows.append(resolved)
    ok = sum(1 for r in rows if r.get("status") == "ok")
    abolished = sum(1 for r in rows if r.get("status") == "abolished")
    return {
        "ok": True,
        "items": rows,
        "summary": {
            "total": len(rows),
            "success": ok,
            "failed": len(rows) - ok,
            "abolished": abolished,
        },
    }
